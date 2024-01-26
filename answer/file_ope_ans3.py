"""
問題3:
Excelファイル grades.xlsx には以下のようなデータが格納されています。
| Name  | Math | English | Science |
|-------|------|---------|---------|
| John  | 90   | 85      | 92      |
| Alice | 78   | 92      | 88      |
| Bob   | 89   | 78      | 90      |

このデータを読み取り、各生徒の科目ごとの平均点を計算して、新しいExcelファイル average_grades.xlsx に出力してください。
"""
import openpyxl

workbook = openpyxl.load_workbook('grades.xlsx')
sheet = workbook.active

averages = {}

for row in sheet.iter_rows(min_row=2, values_only=True):
    student_name, math, english, science = row
    averages[student_name] = {'Math': math, 'English': english, 'Science': science}

for subject in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
    if subject[0] != 'Name':
        subject_name = subject[0]
        subject_averages = sum(averages[student_name][subject_name] for student_name in averages) / len(averages)
        sheet.append([subject_name, subject_averages])

workbook.save('average_grades.xlsx')