"""
問題8:
Excelファイル employees.xlsx には以下のようなデータが格納されています。

| ID  | Name    | Salary |
|-----|---------|--------|
| 101 | John    | 50000  |
| 102 | Alice   | 60000  |
| 103 | Bob     | 55000  |

このデータを読み取り、CSV形式に変換して、新しいCSVファイル employees.csv に保存してください。
"""
import openpyxl
import csv

workbook = openpyxl.load_workbook('employees.xlsx')
sheet = workbook.active

with open('employees.csv', 'w', newline='') as csv_file:
    csv_writer = csv.writer(csv_file)
    csv_writer.writerow(sheet[1][i].value for i in range(1, sheet.max_column + 1))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        csv_writer.writerow(row)