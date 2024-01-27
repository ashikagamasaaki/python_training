"""
問題6:
Excelファイル inventory.xlsx には以下のようなデータが格納されています。

| Product | Stock |
|---------|-------|
| Apple   | 15    |
| Orange  | 8     |
| Banana  | 5     |
| Grape   | 12    |

各商品の在庫数が10以下の場合、その商品名と在庫数を表示するプログラムを作成してください。
"""
import openpyxl

# workbook = openpyxl.Workbook()
# sheet = workbook["Sheet"]

# sheet.append({1:'Product', 2:'Stock'})
# sheet.append({1:'Apple', 2:15})
# sheet.append({1:'Orange', 2:8})
# sheet.append({1:'Banana', 2:5})
# sheet.append({1:'Grape', 2:12})

# workbook.save('resources/inventory.xlsx')

workbook = openpyxl.load_workbook('resources/inventory.xlsx')
sheet = workbook["Sheet"]

for row in sheet.iter_rows(min_row=2,values_only=True):
    product, stock = row
    if stock <= 10:
        print(f'{product}の在庫が{stock}個です。発注してください。')
