"""
問題3:
Excelファイル grades.xlsx には以下のようなデータが格納されています。
| Name  | Math | English | Science |
|-------|------|---------|---------|
| John  | 90   | 85      | 92      |
| Alice | 78   | 92      | 88      |
| Bob   | 89   | 78      | 90      |

このデータを読み取り、各生徒の科目ごとの平均点を計算して、新しいExcelファイル average_grades.xlsx に出力してください。
"""
import openpyxl

# 一旦エクセル作成
# datas = [['Name', 'Math', 'English', 'Science'], ['John', 90, 85, 92], ['Alice', 78, 92, 88], ['Bob', 89, 78, 90]]

# wbname = 'resources/grades.xlsx'
# wb = openpyxl.Workbook()    #　エクセルブックOPEN
# st = wb["Sheet"]    # シート取得

# # a1 = st["A1"]
# # b1 = st.cell(1, 2)
# # a1.value = "Name"
# # b1.value = "Math"

# for r, data in enumerate(datas, 1):
#     for c, val in enumerate(data, 1):
#         cl = st.cell(r, c)
#         cl.value = val

# wb.save(wbname)     # エクセル保存


workbook = openpyxl.load_workbook('resources/grades.xlsx')   #　エクセルブックOPEN
st = workbook.active

average = {}

for row in st.iter_rows(min_row=2, values_only=True):
    name, math, english, science = row
    average[name] = {'Name': name, 'Math': math, 'English': english, 'Science': science}
    print(f'name={name}, math={math}, english={english}, science={science}')

append_dict = {}

for index,subject in enumerate(st.iter_cols(min_row=1, max_row=1, values_only=True), 1):
    if subject[0] != 'Name':
        subject_name = subject[0]
        subject_average = sum(average[student_name][subject_name] for student_name in average) // len(average)
        append_dict[index] = subject_average
    else:
        append_dict[index] = 'Average'

print('append_dict = ', append_dict)
st.append(append_dict)
workbook.save('resources/average_grades.xlsx')





# 読み込み→新規エクセル作成処理
# wb_list = []

# for index, row in enumerate(st.iter_rows(min_row=1, min_col=1)):
#     if index == 0:
#         row.append('Average')
#     value_list = []
#     for c in row:
#         if index == 0:
#             c.append('Average')
#         else:
#             point_list = c.value[1:]
#             point_avg = sum(point_list) / len(point_list)
#             c.value.append(point_avg)
#         value_list.append(c.value)
#     wb_list.append(value_list)


# new_wb_name = 'resources/average_grades.xlsx'
# new_wb = openpyxl.Workbook()
# new_st = new_wb["average"]

# for row, datas in enumerate(wb_list):
#     for col, data in enumerate(datas):
#         new_cl = new_st.cell(row, col)
#         new_cl = data

# new_wb.save(new_wb_name)
