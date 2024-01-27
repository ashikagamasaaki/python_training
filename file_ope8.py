"""
問題8:
Excelファイル employees.xlsx には以下のようなデータが格納されています。

| ID  | Name    | Salary |
|-----|---------|--------|
| 101 | John    | 50000  |
| 102 | Alice   | 60000  |
| 103 | Bob     | 55000  |

このデータを読み取り、CSV形式に変換して、新しいCSVファイル employees.csv に保存してください。
"""
import openpyxl
import csv
# workbook = openpyxl.Workbook()
# sheet = workbook["Sheet"]

# sheet.cell(1,1).value = 'ID'
# sheet.cell(1,2).value = 'Name'
# sheet.cell(1,3).value = 'Salary'

# sheet.append({1:101, 2:'John', 3:50000})
# sheet.append({1:102, 2:'Alice', 3:60000})
# sheet.append({1:103, 2:'Bob', 3:55000})

# workbook.save('resources/employees.xlsx')

workbook = openpyxl.load_workbook('resources/employees.xlsx')
sheet = workbook["Sheet"]

title = []

for col in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
    title.append(col[0])


with open('resources/employees.csv', 'w', encoding='utf-8', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=title)
    writer.writeheader()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        id, name, salary = row
        row_dict = {key: value for key, value in zip(title, list(row))}
        writer.writerow(row_dict)
